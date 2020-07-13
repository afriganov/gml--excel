from lxml import etree
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()
file_path = filedialog.askdirectory()

class Dejtafrejm:
    def __init__(self,put,*argv):
        self.put = put
        self.tree = etree.parse(file_path+"/"+self.put)
        self.root = self.tree.getroot()
        self.data = {}
        for arg in argv:
            setattr(self, "lista_"+arg.lower()+"_"+self.put[:-4].lower(), [])
            for a in self.root.iter("{http://www.uredjenazemlja.hr}"+arg):
                getattr(self, "lista_"+arg.lower()+"_"+self.put[:-4].lower()).append(a.text)
            b = {arg.replace("_"," ").lower().capitalize():getattr(self, "lista_" + arg.lower() + "_" + self.put[:-4].lower())}
            self.data.update(b)
        self.dataframe = pd.DataFrame(self.data).fillna("")

a = Dejtafrejm("NACINI_UPORABE.gml", "POSJEDOVNI_LIST_BROJ", "POVRSINA", "CESTICA_ID", "NAZIV_VRSTE_UPORABE")
b = Dejtafrejm("CESTICE.gml", "CESTICA_ID", "BROJ_CESTICE", "ADRESA_OPISNA")
c = Dejtafrejm("POSJEDOVNI_LISTOVI.gml", "POSJEDOVNI_LIST_ID", "BROJ")
d = Dejtafrejm("UPISANE_OSOBE.gml", "POSJEDNIK_ID", "POSJEDOVNI_LIST_ID", "POSJEDNIK_ID","POSJEDOVNI_LIST_ID", "POSJEDNIK_UDIO_BROJNIK","POSJEDNIK_UDIO_NAZIVNIK","NAZIV","NAZIV_PRAVNOG_ODNOSA","POREZNI_BROJ")
e = Dejtafrejm("CESTICE_ZGRADE.gml", "CESTICA_ID", "ZGRADA_ID")
f = Dejtafrejm("ZGRADE.gml","ZGRADA_ID","POVRSINA","NAZIV_VRSTE_ZGRADE","POSJEDOVNI_LIST_BROJ")
g = Dejtafrejm("ADRESE_UPISANIH_OSOBA.gml","POSJEDNIK_ID", "DRZAVA", "ADRESA_OPISNA", "POSTANSKI_BROJ", "NASELJE", "ULICA", "KBR")
azk = Dejtafrejm("ZK_CESTICE.gml", "ZK_CESTICA_ID","BROJ_CESTICE","PODBROJ_CESTICE", "ULOZAK_ID", "POVRSINA", "RBR_ZK_TIJELA")
bzk = Dejtafrejm("ULOSCI.gml", "ULOZAK_ID", "BROJ_ULOSKA")
czk = Dejtafrejm("ZK_NACINI_UPORABE.gml", "ZK_CESTICA_ID", "NAZIV_VRSTE_UPORABE", "POVRSINA")
zzk = Dejtafrejm("ZK_ADRESE_CESTICE.gml", "ADRESA_OPISNA", "ZK_CESTICA_ID")
dzk = Dejtafrejm("ZK_VLASNICI.gml", "ULOZAK_ID", "OSOBA_UPISA_ID", "UDIO_DIJELA_BROJNIK", "UDIO_DIJELA_NAZIVNIK", "NAZIV", "ULOZAK_ID", "POREZNI_BROJ", "UDIO_BROJNIK", "UDIO_NAZIVNIK", "RBR_ETAZE", "RBR_ZK_TIJELA")
ezk = Dejtafrejm("ZK_ADRESE_VLASNIKA.gml", "OSOBA_UPISA_ID", "ADRESA_OPISNA")
eaz = Dejtafrejm("ADRESE_ZGRADE.gml", "ZGRADA_ID", "ADRESA_OPISNA")
xzk = Dejtafrejm("ZK_ZGRADE.gml", "POVRSINA", "ZK_CESTICA_ID", "NAZIV_VRSTE_ZGRADE")

#TRENUTNO RJESENJE ! ! ! treba srediti imena (nastalo kod prebacivanja na klasu umjesto špageta)
b.dataframe = b.dataframe.rename(columns={"Adresa opisna": "Adresa"})
c.dataframe = c.dataframe.rename(columns={"Broj": "Posjedovni list broj"})
g.dataframe = g.dataframe.rename(columns={"Kbr": "Kucni broj"})
azk.dataframe = azk.dataframe.rename(columns={"Broj cestice": "Broj cestice brojnik", "Rbr zk tijela":"Redni broj ZK tijela cestice"}) #ovaj brojnik pogotovo nema smisla ?!
czk.dataframe = czk.dataframe.rename(columns={"Povrsina": "Povrsina vrste uporabe"})
dzk.dataframe = dzk.dataframe.rename(columns={"Rbr etaze": "Redni broj etaze", "Rbr zk tijela": "Redni broj ZK tijela vlasnika"})
ezk.dataframe = ezk.dataframe.rename(columns={"Adresa opisna": "Adresa"})
xzk.dataframe = xzk.dataframe.rename(columns={"Povrsina": "Povrsina zgrade"})


d.dataframe = d.dataframe.merge(g.dataframe, on="Posjednik id", how="left")
d.dataframe = d.dataframe.reset_index()


d.dataframe = d.dataframe.merge(c.dataframe, on="Posjedovni list id", how = "left")

d.dataframe['Omjer'] = ""
d.dataframe["Sklopljeno"]=""
d.dataframe["Adresa"]=""

for i in range(0,len(d.dataframe["Posjedovni list id"])):
    d.dataframe.at[i, "Omjer"] = str(d.dataframe.at[i, "Posjednik udio brojnik"]) + "/" + str(d.dataframe.at[i, "Posjednik udio nazivnik"])
    if d.dataframe.at[i,"Adresa opisna"] != None and d.dataframe.at[i,"Naselje"] == None:
        d.dataframe.at[i,"Adresa"] = d.dataframe.at[i,"Adresa opisna"]
    elif d.dataframe.at[i,"Adresa opisna"] == None and d.dataframe.at[i,"Naselje"] != None and d.dataframe.at[i,"Postanski broj"] == None:
        d.dataframe.at[i, "Adresa"] = d.dataframe.at[i, "Ulica"]+" "+str(d.dataframe.at[i, "Kucni broj"]) + ", " + d.dataframe.at[i, "Naselje"]+", "+d.dataframe.at[i, "Drzava"]
    elif d.dataframe.at[i,"Adresa opisna"] == None and d.dataframe.at[i,"Naselje"] != None:
        d.dataframe.at[i, "Adresa"] = d.dataframe.at[i, "Ulica"]+" "+str(d.dataframe.at[i, "Kucni broj"]) + ", " +d.dataframe.at[i, "Naselje"]+", "+d.dataframe.at[i, "Drzava"]
    else:
        pass

for i in range(0,len(d.dataframe["Posjedovni list id"])):
    if str(d.dataframe.at[i,"Naziv pravnog odnosa"]) == "VLASNIK" and d.dataframe.at[i,"Porezni broj"] != "":
        d.dataframe.at[i,"Sklopljeno"] =d.dataframe.at[i,"Naziv"]+", "+d.dataframe.at[i, "Adresa"]+", (Vlasnik), "+"OIB: "+str(d.dataframe.at[i,"Porezni broj"])
    elif str(d.dataframe.at[i,"Naziv pravnog odnosa"]) == "VLASNIK" and d.dataframe.at[i,"Porezni broj"] == "" and d.dataframe.at[i,"Adresa"] != "":
        d.dataframe.at[i,"Sklopljeno"] = d.dataframe.at[i,"Naziv"]+", "+d.dataframe.at[i, "Adresa"]+", (Vlasnik)"
    elif str(d.dataframe.at[i,"Naziv pravnog odnosa"]) == "VLASNIK" and d.dataframe.at[i,"Porezni broj"] == "" and d.dataframe.at[i,"Adresa"] == "":
        d.dataframe.at[i,"Sklopljeno"] = d.dataframe.at[i,"Naziv"]+", (Vlasnik)"
    elif d.dataframe.at[i,"Porezni broj"] == "":
        d.dataframe.at[i, "Sklopljeno"] = d.dataframe.at[i, "Naziv"] + ", " + d.dataframe.at[i, "Adresa"]
    else:
        d.dataframe.at[i, "Sklopljeno"] = d.dataframe.at[i, "Naziv"] + ", " + d.dataframe.at[i, "Adresa"] + ", " + "OIB: " + str(d.dataframe.at[i, "Porezni broj"])

dataframeac = e.dataframe.merge(f.dataframe,on="Zgrada id",how="left")
dataframeac = dataframeac.merge(eaz.dataframe,on="Zgrada id",how="left")
dataframeac["Naziv vrste uporabe"]=""
dataframeac = dataframeac.fillna("")

for i in range(0,len(dataframeac["Cestica id"])):
    if dataframeac.at[i, "Adresa opisna"] != "":
        dataframeac.at[i, "Naziv vrste uporabe"] = dataframeac.at[i, "Naziv vrste zgrade"]+", "+dataframeac.at[i, "Adresa opisna"]
    else:
        dataframeac.at[i, "Naziv vrste uporabe"] = dataframeac.at[i, "Naziv vrste zgrade"]

dataframeac = dataframeac.merge(b.dataframe,on="Cestica id",how="left")
dataframeac.to_excel(file_path+"/dataframeac.xlsx")
dataframeac.drop(columns="Adresa opisna")
dataframeac["Adresa"]=""
dataframeac = dataframeac[["Posjedovni list broj","Cestica id","Naziv vrste uporabe","Povrsina","Broj cestice","Adresa"]]

dataframeab = a.dataframe.merge(b.dataframe,on="Cestica id",how="left")
dataframeab = dataframeab.append(dataframeac,ignore_index=True)
dataframeab = dataframeab.sort_values(["Posjedovni list broj","Broj cestice"],ignore_index=True)

i = 0
j = 1
for z in range(0,len(dataframeab["Posjedovni list broj"])-1):
    if dataframeab.at[i,"Posjedovni list broj"] == dataframeab.at[i+j,"Posjedovni list broj"] and dataframeab.at[i,"Broj cestice"] == dataframeab.at[i+j,"Broj cestice"]:
        dataframeab.at[i + j, "Posjedovni list broj"] = ""
        dataframeab.at[i + j, "Adresa"] = ""
        j = j + 1
        i = i
    else:
        i = i + j
        j = 1


dataframeab = dataframeab.merge(d.dataframe, on="Posjedovni list broj", how = "left")
dataframeab = dataframeab.sort_values(["Broj cestice","Posjedovni list broj","Naziv vrste uporabe"],ignore_index=True,ascending=[True,False,True])
dataframeab = dataframeab[["Posjedovni list broj","Broj cestice","Naziv vrste uporabe","Omjer","Sklopljeno","Povrsina","Adresa_x"]]

i = 0
j = 1
for z in range(0,len(dataframeab["Posjedovni list broj"])-1):
    if dataframeab.at[i,"Posjedovni list broj"] == dataframeab.at[i+j,"Posjedovni list broj"] and dataframeab.at[i,"Broj cestice"] == dataframeab.at[i+j,"Broj cestice"] and dataframeab.at[i, "Naziv vrste uporabe"] == dataframeab.at[i + j, "Naziv vrste uporabe"] and dataframeab.at[i, "Povrsina"] == dataframeab.at[i + j, "Povrsina"] and dataframeab.at[i, "Adresa_x"] == dataframeab.at[i + j, "Adresa_x"]:
        dataframeab.at[i + j, "Posjedovni list broj"] = ""
        dataframeab.at[i + j, "Broj cestice"] = ""
        dataframeab.at[i + j, "Naziv vrste uporabe"] = ""
        dataframeab.at[i + j, "Povrsina"] = ""
        dataframeab.at[i + j, "Adresa_x"] = ""
        j = j + 1
        i = i
    else:
        i = i + j
        j = 1

for z in range(0,len(dataframeab["Posjedovni list broj"])-1):
    if dataframeab.at[z,"Posjedovni list broj"] == "":
        dataframeab.at[z, "Broj cestice"] = ""

for z in range(0,10):
    for i in range(0,len(dataframeab["Posjedovni list broj"])-1):
        if dataframeab.at[i,"Posjedovni list broj"] == "" and dataframeab.at[i, "Naziv vrste uporabe"] != "" and dataframeab.at[i, "Povrsina"] != "" and dataframeab.at[i-1, "Naziv vrste uporabe"] == "":
            dataframeab.at[i-1, "Naziv vrste uporabe"] = dataframeab.at[i, "Naziv vrste uporabe"]
            dataframeab.at[i-1, "Povrsina"] = dataframeab.at[i, "Povrsina"]
            dataframeab.at[i, "Povrsina"] = ""
            dataframeab.at[i, "Naziv vrste uporabe"] = ""
        else:
            pass

dataframeab = dataframeab.fillna("")
dataframeab["Drop"] = None

for i in range(0,len(dataframeab["Posjedovni list broj"])):
    if dataframeab.at[i,"Posjedovni list broj"] == "" and dataframeab.at[i, "Broj cestice"] == "" and dataframeab.at[i, "Naziv vrste uporabe"] == "" and dataframeab.at[i, "Omjer"] == "" and dataframeab.at[i, "Sklopljeno"] == "" and dataframeab.at[i, "Povrsina"] == "" and dataframeab.at[i, "Adresa_x"] == "":
        dataframeab.at[i,"Drop"] = None
    else:
        dataframeab.at[i,"Drop"] = "1"

dataframeab = dataframeab.dropna(subset=["Drop"])

#######################################################################################################################################################

xzk.dataframe = xzk.dataframe.merge(azk.dataframe, on="Zk cestica id", how="left")
xzk.dataframe = xzk.dataframe[["Zk cestica id","Broj cestice brojnik","Podbroj cestice","Ulozak id","Povrsina zgrade","Redni broj ZK tijela cestice","Naziv vrste zgrade"]]
azk.dataframe = azk.dataframe.append(xzk.dataframe,ignore_index=True)
azk.dataframe = azk.dataframe.merge(zzk.dataframe, on="Zk cestica id", how="left")
dataframeabzk = azk.dataframe.merge(bzk.dataframe, on="Ulozak id", how="left")
dataframeabzk = dataframeabzk.merge(czk.dataframe, on="Zk cestica id", how="outer")
dzk.dataframe = dzk.dataframe.merge(ezk.dataframe, on= "Osoba upisa id", how = "left")


dzk.dataframe["Nazadr"] =""
dzk.dataframe["Omjer"] =""
dzk.dataframe = dzk.dataframe.fillna("")

for i in range(0,len(dzk.dataframe["Osoba upisa id"])):
    dzk.dataframe.at[i, "Omjer"] = str(dzk.dataframe.at[i, "Udio brojnik"]) + "/" + str(dzk.dataframe.at[i, "Udio nazivnik"])
    if dzk.dataframe.at[i,"Porezni broj"] != "":
        dzk.dataframe.at[i,"Nazadr"] = (dzk.dataframe.at[i,"Naziv"]+", "+ "OIB: "+dzk.dataframe.at[i,"Porezni broj"]+", "+str(dzk.dataframe.at[i,"Adresa"])).upper()
    elif dzk.dataframe.at[i,"Porezni broj"] == "" and dzk.dataframe.at[i, "Adresa"] != "":
        dzk.dataframe.at[i, "Nazadr"] = (dzk.dataframe.at[i, "Naziv"] +", "+ str(dzk.dataframe.at[i, "Adresa"])).upper()
    else:
        dzk.dataframe.at[i, "Nazadr"] = str(dzk.dataframe.at[i, "Naziv"]).upper()

for i in range(0,len(dzk.dataframe["Osoba upisa id"])):
    if  dzk.dataframe.at[i,"Redni broj etaze"] != "" and dzk.dataframe.at[i,"Udio dijela brojnik"] == "0" and dzk.dataframe.at[i,"Udio dijela nazivnik"] == "0":
        dzk.dataframe.at[i, "Naziv"] = dzk.dataframe.at[i, "Omjer"] +" "+ dzk.dataframe.at[i, "Nazadr"]
        dzk.dataframe.at[i,"Omjer"] = "(E-"+dzk.dataframe.at[i,"Redni broj etaze"]+") NEODREĐEN OMJER"
    elif dzk.dataframe.at[i,"Redni broj etaze"] != "" and dzk.dataframe.at[i,"Udio dijela brojnik"] != "":
        dzk.dataframe.at[i, "Naziv"] = dzk.dataframe.at[i, "Omjer"] +" "+ dzk.dataframe.at[i, "Nazadr"]
        dzk.dataframe.at[i, "Omjer"] = dzk.dataframe.at[i,"Udio dijela brojnik"]+"/"+dzk.dataframe.at[i,"Udio dijela nazivnik"]+" ETAŽNO VLASNIŠTVO (E-"+dzk.dataframe.at[i,"Redni broj etaze"]+")"
    else:
        dzk.dataframe.at[i,"Omjer"] = dzk.dataframe.at[i,"Udio dijela brojnik"]+"/"+dzk.dataframe.at[i,"Udio dijela nazivnik"]
        dzk.dataframe.at[i,"Naziv"] =  dzk.dataframe.at[i,"Nazadr"]

dzk.dataframe.to_excel(file_path+"/outputfinalnidzk.xlsx")
dzk.dataframe = dzk.dataframe.sort_values(["Redni broj ZK tijela vlasnika"],ignore_index=True)
dataframeabzk=dataframeabzk.fillna("")

for i in range(0,len(dataframeabzk["Zk cestica id"])):
    if dataframeabzk.at[i,"Podbroj cestice"] != "":
        dataframeabzk.at[i,"Broj cestice"] = dataframeabzk.at[i,"Broj cestice brojnik"]+"/"+dataframeabzk.at[i,"Podbroj cestice"]
    else:
        dataframeabzk.at[i, "Broj cestice"] = dataframeabzk.at[i, "Broj cestice brojnik"]
for i in range(0,len(dataframeabzk["Broj cestice"])):
    if dataframeabzk.at[i,"Povrsina zgrade"] != "" and dataframeabzk.at[i,"Naziv vrste zgrade"] != "":
        dataframeabzk.at[i,"Povrsina vrste uporabe"] = dataframeabzk.at[i,"Povrsina zgrade"]
        dataframeabzk.at[i,"Naziv vrste uporabe"] = dataframeabzk.at[i,"Naziv vrste zgrade"]
for i in range(0,len(dataframeabzk["Broj cestice"])):
    if dataframeabzk.at[i, "Broj cestice"].endswith("ZGR"):
        dataframeabzk.at[i, "Broj cestice"] = "*" + str(dataframeabzk.at[i, "Broj cestice"])[:-4]
    elif dataframeabzk.at[i, "Broj cestice"].endswith("zgr"):
        dataframeabzk.at[i, "Broj cestice"] = "*" + str(dataframeabzk.at[i, "Broj cestice"])[:-4]
    else:
        dataframeabzk.at[i, "Broj cestice"] = str(dataframeabzk.at[i, "Broj cestice"])

dataframeabzk = dataframeabzk.sort_values(["Broj cestice","Broj uloska","Naziv vrste uporabe"],ignore_index=True)

i = 0
j = 1
for z in range(0,len(dataframeabzk["Ulozak id"])-1):
    if dataframeabzk.at[i,"Broj cestice"] == dataframeabzk.at[i+j,"Broj cestice"] and dataframeabzk.at[i,"Broj uloska"] == dataframeabzk.at[i+j,"Broj uloska"]:
        dataframeabzk.at[i+j, "Ulozak id"] = ""
        dataframeabzk.at[i + j, "Broj cestice"] = ""
        dataframeabzk.at[i+j,"Broj uloska"] = ""
        j = j + 1
    else:
        i = i + j
        j = 1


dataframeabzk = dataframeabzk.merge(dzk.dataframe, on="Ulozak id", how="left")
dataframeabzk = dataframeabzk[["Broj uloska","Broj cestice","Omjer","Naziv","Naziv vrste uporabe","Povrsina vrste uporabe","Povrsina","Adresa opisna","Redni broj ZK tijela cestice","Redni broj ZK tijela vlasnika"]]

i = 0
j = 1
for z in range(0,len(dataframeabzk["Broj uloska"])-1):
    if dataframeabzk.at[i,"Broj cestice"] == dataframeabzk.at[i+j,"Broj cestice"] and dataframeabzk.at[i,"Broj uloska"] == dataframeabzk.at[i+j,"Broj uloska"]:
        dataframeabzk.at[i + j, "Broj cestice"] = ""
        dataframeabzk.at[i+j,"Broj uloska"] = ""
        dataframeabzk.at[i+j,"Povrsina"] = ""
        j = j + 1
    else:
        i = i + j
        j = 1


i = 0
j = 1
for z in range(0,len(dataframeabzk["Broj uloska"])-1):
    if dataframeabzk.at[i,"Naziv vrste uporabe"] == dataframeabzk.at[i+j,"Naziv vrste uporabe"] and dataframeabzk.at[i,"Povrsina vrste uporabe"] == dataframeabzk.at[i+j,"Povrsina vrste uporabe"]:
        dataframeabzk.at[i+j, "Naziv vrste uporabe"] = ""
        dataframeabzk.at[i+j,"Povrsina vrste uporabe"] = ""
        j = j + 1
    else:
        i = i + j
        j = 1

i = 0
j = 1
for z in range(0,len(dataframeabzk["Broj uloska"])-1):
    if dataframeabzk.at[i,"Adresa opisna"] == dataframeabzk.at[i+j,"Adresa opisna"] and dataframeabzk.at[i+j,"Broj cestice"] == "":
        dataframeabzk.at[i+j, "Adresa opisna"] = ""
        j = j + 1
    else:
        i = i + j
        j = 1

dataframeabzk = dataframeabzk.fillna("")
for i in range(0,len(dataframeabzk["Broj uloska"])-1):
    if dataframeabzk.at[i,"Naziv vrste uporabe"] != "" and dataframeabzk.at[i,"Naziv"] == "":
        dataframeabzk.at[i, "Povrsina"] = ""
        i = i + 1
    else:
        pass

for i in range(0,len(dataframeabzk["Broj uloska"])-1):
    if dataframeabzk.at[i,"Broj cestice"] == "":
        dataframeabzk.at[i, "Redni broj ZK tijela cestice"] = ""
        i = i + 1
    else:
        pass

for z in range(0,50):
    for i in range(0,len(dataframeabzk["Broj uloska"])-1):
        if dataframeabzk.at[i,"Broj uloska"] == "" and dataframeabzk.at[i, "Naziv vrste uporabe"] != "" and dataframeabzk.at[i, "Povrsina vrste uporabe"] != "" and dataframeabzk.at[i-1, "Naziv vrste uporabe"] == "":
            dataframeabzk.at[i-1, "Naziv vrste uporabe"] = dataframeabzk.at[i, "Naziv vrste uporabe"]
            dataframeabzk.at[i-1, "Povrsina vrste uporabe"] = dataframeabzk.at[i, "Povrsina vrste uporabe"]
            dataframeabzk.at[i, "Povrsina vrste uporabe"] = ""
            dataframeabzk.at[i, "Naziv vrste uporabe"] = ""
        else:
            pass

dataframeabzk["Drop"] = None
for i in range(0,len(dataframeabzk["Broj uloska"])):
    if dataframeabzk.at[i,"Broj uloska"] == "" and dataframeabzk.at[i, "Broj cestice"] == "" and dataframeabzk.at[i, "Omjer"] == "" and dataframeabzk.at[i, "Naziv"] == "" and dataframeabzk.at[i, "Naziv vrste uporabe"] == "" and dataframeabzk.at[i, "Povrsina vrste uporabe"] == "" and dataframeabzk.at[i, "Povrsina"] == "":
        dataframeabzk.at[i,"Drop"] = None
    else:
        dataframeabzk.at[i,"Drop"] = "1"
dataframeabzk = dataframeabzk.dropna(subset=["Drop"])
cestcut = pd.DataFrame()
cestcut = dataframeab[["Broj cestice"]].copy()
cestcut.to_excel(file_path+"/cestcut.xlsx")
cestzk = []
j = 1


################################## Ako treba nešto izbaciti
#dataframeabzk = dataframeabzk.reset_index(drop = True)
#dataframeabzk = dataframeabzk.drop([58,59,60,61,62])
##################################


dataframeabzk = dataframeabzk.reset_index(drop=True)
for i in range(0,len(dataframeabzk["Broj cestice"])):
    if dataframeabzk.at[i,"Broj cestice"] != "":
        dataframeabzk.at[i,"Drop"] = j
        j = j+1
        cestzk.append(dataframeabzk.at[i,"Broj cestice"])
    else:
        dataframeabzk.at[i,"Drop"] = ""

for i in range(0,len(dataframeabzk["Broj cestice"])):
    if dataframeabzk.at[i,"Naziv vrste uporabe"] == "":
        dataframeabzk.at[i,"Naziv vrste uporabe"] = dataframeabzk.at[i,"Adresa opisna"]
    else:
        pass
del dataframeabzk["Adresa opisna"]

def duplicipos(i):
    if dataframeab.at[i,"Broj cestice"] in lista:
        return True
    else:
        return False
def cestzkfun(i):
    if dataframeab.at[i,"Broj cestice"] in cestzk:
        return True
    else:
        return False

dataframeab = dataframeab.reset_index(drop = True)
lista=[]

j = 1
for i in range(0,len(dataframeab["Broj cestice"])):
    if dataframeab.at[i,"Broj cestice"] != "" and duplicipos(i) == False and cestzkfun(i) == True:
        dataframeab.at[i,"Drop"] = j
        j = j+1
        lista.append(dataframeab.at[i,"Broj cestice"])
    else:
        dataframeab.at[i,"Drop"] = ""


dataframeabzk.to_excel(file_path+"/outputfinalniabzk.xlsx")
dataframeab.to_excel(file_path+"/outputfinalniab.xlsx")
wbab = load_workbook(file_path+"/outputfinalniab.xlsx")
wbabzk = load_workbook(file_path+"/outputfinalniabzk.xlsx")
wsbab = wbab.active
wsbabzk = wbabzk.active

if len(dataframeab["Broj cestice"]) > len(dataframeabzk["Broj cestice"]):
    a = len(dataframeab["Broj cestice"])
else:
    a = len(dataframeabzk["Broj cestice"])


for i in range(2,a+100):
    if wsbab.cell(row=i, column=9).value != wsbabzk.cell(row=i, column=11).value and wsbab.cell(row=i, column=9).value == None and wsbabzk.cell(row=i, column=11).value != None:
        wsbabzk.insert_rows(i)
        i = i
    if wsbab.cell(row=i, column=9).value != wsbabzk.cell(row=i, column=11).value and wsbab.cell(row=i, column=9).value != None and wsbabzk.cell(row=i, column=11).value == None:
        wsbab.insert_rows(i)
        i = i
    else:
        i = i + 1

masterpos = pd.DataFrame(wsbab.values)
masterzk = pd.DataFrame(wsbabzk.values)
masterpos.to_excel(file_path+"/masterpos.xlsx")
masterzk.to_excel(file_path+"/masterzk.xlsx", index=False)
masterpos.columns = masterpos.iloc[0]

masterpos = masterpos.drop(masterpos.index[0])
masterpos = masterpos.set_index(["Drop"])
masterpos.columns = ["A","Broj posjedovnog lista","Broj cestice","Način uporabe katastarske čestice","Omjer","Prezime i ime odnosno tvrtka ili naziv upisane osobe.Prebivalište odnosno sjedište, ulica , kućni broj i OIB upisane osobe (posjednika).","Povrsina [m2]","Adresa"]
del masterpos["A"]

masterzk.columns = masterzk.iloc[0]
masterzk = masterzk.drop(masterzk.index[0])
masterzk.columns = ["A","Broj Z.K. uloška","Broj katastarske čestice","Omjer","Prezime i ime odnosno tvrtka ili naziv upisane osobe.Prebivalište odnosno sjedište, ulica , kućni broj i OIB upisane osobe (vlasnika).","Način uporabe katastarske čestice","Površina vrste uporabe", "Površina ukupna [m2]","Redni broj ZK tijela cestice","Redni broj ZK tijela vlasnika","Drop"]
del masterzk["A"]
del masterzk["Drop"]


masterpos.to_excel(file_path+"/masterpos.xlsx")
masterzk.to_excel(file_path+"/masterzk.xlsx", index=False)
ws = load_workbook(file_path+"/masterpos.xlsx")
wss = load_workbook(file_path+"/masterzk.xlsx")
wsa = ws.active
wsb = wss.active
wsa.column_dimensions['A'].width = 9
wsa.column_dimensions['B'].width = 13
wsa.column_dimensions['C'].width = 13
wsa.column_dimensions['D'].width = 13
wsa.column_dimensions['E'].width = 10
wsa.column_dimensions['F'].width = 50
wsa.column_dimensions['G'].width = 13
wsa.column_dimensions['H'].width = 13

for row in wsa.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True, horizontal="center", vertical="center")

wsb.column_dimensions['A'].width = 13
wsb.column_dimensions['B'].width = 13
wsb.column_dimensions['C'].width = 10
wsb.column_dimensions['D'].width = 50
wsb.column_dimensions['E'].width = 13
wsb.column_dimensions['F'].width = 13
wsb.column_dimensions['G'].width = 13
for row in wsb.iter_rows():
    for cell in row:
        cell.alignment = cell.alignment.copy(wrapText=True, horizontal="center", vertical="center")
ws.save(file_path+"/masterpos.xlsx")
wss.save(file_path+"/masterzk.xlsx")





